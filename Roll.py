from openpyxl import load_workbook
from openpyxl.styles import Font
import win32com.client as win32
from datetime import datetime
import os


def roll_no():
	try:
		os.listdir(f'D:\\Piyush\\Roll\\{DATE.year}\\{DATE.month}')
	except FileNotFoundError:
		os.mkdir(f'D:\\Piyush\\Roll\\{DATE.year}\\{DATE.month}')

	for i in range(1,31):
		if f'{i}.xlsx' not in os.listdir(f'D:\\Piyush\\Roll\\{DATE.year}\\{DATE.month}'):
			return i


def cnvrt():
	excel = win32.Dispatch('Excel.Application')#excel = win32.gencache.EnsureDispatch('Excel.Application')
	wb = excel.Workbooks.Open(DIR)
	wb.SaveAs(DIR, FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
	wb.Close()                         #FileFormat = 56 is for .xls extension
	excel.Application.Quit()


def book_count():
	for i in range(14, 50):
		if(rollsht.cell(row=i, column=3).value == "E-Banking Ref No"):
			return i-14


def move_over():

	#---Moving Everything
	for i in range(2, mc):
		for j in range(14, mr+1):
			try:
				tmplsht.cell(row=j, column=i).value = rollsht.cell(row=j, column=i).value
			except AttributeError:
				continue
	
	#---Adding series
	for i in range(14, mr+1):
		tmplsht.cell(row=i, column=2).value = i-13

	#---Moving Misc. Stuff
	tmplsht.cell(row=5, column=9).value = rollsht.cell(row=5, column=9).value
	tmplsht.cell(row=6, column=9).value = rollsht.cell(row=6, column=9).value
	tmplsht.cell(row=45, column=3).value = rollsht.cell(row=mr+2, column=3).value
	tmplsht.cell(row=45, column=10).value = rollsht.cell(row=mr+2, column=10).value


def re_size():
	style = Font(size=8, bold=False)
	bdstyle = Font(size=16, bold=True)

	#---EVERYTHING...
	for i in range(1, mc):
		for j in range(1, 44):
			tmplsht.cell(row=j, column=i).font = style
	
	#---Bottom stuff
	tmplsht.cell(row=45, column=3).font = bdstyle
	tmplsht.cell(row=45, column=10).font = bdstyle
	
	#---Top stuff
	for i in range(4, 8):
		tmplsht.cell(row=i, column=9).font = bdstyle

	for i in range(4, 8):
		tmplsht.cell(row=i, column=9).font = bdstyle

	for i in range(14, mr+1):
		tmplsht.cell(row=i, column=7).font = Font(name='Candara', size=9, bold=True)


def save():
	#tmpl.save(f'{ROLLPATH}{roll_num}.xlsx') 
	tmpl.save(f'D:\\Piyush\\Roll\\{DATE.year}\\{DATE.month}\\{roll_num}.xlsx')
	print(f'{file} Saved as {roll_num}')


def make_record():
	try:
		rcrd = load_workbook(f'D:\\Piyush\\Roll\\{DATE.year}\\{DATE.month}\\Records.xlsx')
	except FileNotFoundError:
		rcrd = load_workbook(f'D:\\Piyush\\Roll\\Record_template.xlsx')
	rcrdsht = rcrd.active

	for r in range(2, 30):
		if rcrdsht.cell(row=r, column=1).value == None:	
			rcrdsht.cell(row=r, column=1).value = roll_no() - 1
			rcrdsht.cell(row=r, column=2).value = tmplsht.cell(row=6, column=9).value
			rcrdsht.cell(row=r, column=3).value = book_count()
			rcrdsht.cell(row=r, column=4).value = tmplsht.cell(row=45, column=10).value
			break
	
	rcrd.save(f'D:\\Piyush\\Roll\\{DATE.year}\\{DATE.month}\\Records.xlsx')


def clnup():
	os.remove(f'{DIR}.xlsx') 
	#os.remove(f'{DIR}.xls')


def print_it():
	print_que = input("Enter RollNo:- ")
	if print_que.lower()=="y":
		for _ in range(3):
			os.startfile(f"D:\\Piyush\\Roll\\{DATE.year}\\{DATE.month}\\{roll_num}.xlsx", "print")
A = input("Procced?")    






ROLLPATH = 'C:\\Users\\shara\\Downloads\\'
DATE = datetime.now()

for file in os.listdir(ROLLPATH):
	if file.startswith('RDInstallmentReport') and file.endswith('.xls'):
		DIR = ROLLPATH + file[0:-4]
		roll_num = roll_no()

		cnvrt()

		#---Open Workbooks
		tmpl = load_workbook('D:\\Piyush\\Roll\\RollTamplate.xlsx')
		roll = load_workbook(DIR + '.xlsx')
		tmplsht = tmpl.active
		rollsht = roll.active

		#---Much needed Variables
		bkcount = book_count()
		mc = 24
		mr = bkcount + 13

		#---Work on it & save it.
		move_over()
		re_size()
		save()
		make_record()
		clnup()
		#print_it()
